import pandas as pd
import numpy as np
import re
from openpyxl.utils import range_boundaries

class ExcelSheetParser:
    def __init__(self, sheet, maxn=10, model=None, debug=True):
        self.sheet = sheet
        self.maxn = maxn
        self.model = model
        self.debug = debug
        self.df = None
        self.hasmerged = False
        self.datarow = None
        self.report = None

    def get_sheet_attributes(self):
        """
        Returns a set of table attributes for the given sheet
        """
        dd = pd.DataFrame(self.sheet.values)

        null_cells_in_rows = list(dd[0:self.maxn].apply(lambda x: x.isnull().sum(), axis="columns"))
        float_cells_in_rows = []
        unique_vals_in_rows = []
        year_vals_in_rows = []
        hxl_row = None

        for index, row in dd[0:self.maxn].iterrows():
            unique_vals = list(row.unique())
            unique_vals = [i for i in unique_vals if i is not None and str(i) != "nan"]
            unique_vals_in_rows.append(len(unique_vals))

            float_count = 0
            year_count = 0
            if self.check_hdx_header(list(row)):
                hxl_row = index
            for col in dd.columns:
                val = row[col]
                val = self.process_cell_value(val)
                if self.is_year_value(val):
                    year_count += 1
                elif isinstance(val, (float, int)) or "^=" in str(row[col]):
                    float_count += 1

            float_cells_in_rows.append(float_count)
            year_vals_in_rows.append(year_count)

        first_float_row, first_not_null_row = self.calculate_first_rows(null_cells_in_rows, float_cells_in_rows)

        report = (
            f"Nulls in first {self.maxn} rows: {str(null_cells_in_rows)}\n"
            f"Numeric first {self.maxn} rows: {str(float_cells_in_rows)}\n"
            f"Unique values in first {self.maxn} rows: {str(unique_vals_in_rows)}\n"
            f"Year values in first {self.maxn} rows: {str(year_vals_in_rows)}\n"
            f"HXL row: {str(hxl_row)}\n"
            f"\nFirst reduced nulls row: {str(first_not_null_row)}\n"
            f"First increased numeric row (excluding years): {str(first_float_row)}\n"
        )

        report_json = {
            "null_cells_in_rows": null_cells_in_rows,
            "float_cells_in_rows": float_cells_in_rows,
            "unique_vals_in_rows": unique_vals_in_rows,
            "year_vals_in_rows": year_vals_in_rows,
            "hxl_row": hxl_row,
            "first_float_row": first_float_row,
            "first_not_null_row": first_not_null_row,
        }

        return report, report_json

    def process_cell_value(self, val):
        if isinstance(val, str):
            val = val.replace(",", "").replace(" ", "")
            if val.isnumeric():
                val = int(val)
        return val

    def is_year_value(self, val):
        return isinstance(val, (int, float)) and 1900 < val < 2100

    def calculate_first_rows(self, null_cells_in_rows, float_cells_in_rows):
        first_float_row = 0
        max_floats = max(float_cells_in_rows)
        min_nulls = min(null_cells_in_rows)
        
        for i in range(1, len(float_cells_in_rows)):
            if float_cells_in_rows[i] / max_floats > 0.5 or (
                float_cells_in_rows[i] > 0 and float_cells_in_rows[i - 1] == 0
            ):
                first_float_row = i
                break

        first_not_null_row = np.argmin(null_cells_in_rows)
        return first_float_row, first_not_null_row

    def check_hdx_header(self, first_row):
        matches = ["#meta", "#country", "#data", "#loc", "#geo"]
        first_row = str(first_row)
        return any(x in first_row for x in matches)

    def pad_merged_cells(self):
        """
        Unmerge merged cells and fill with merged value.
        """
        self.df = pd.DataFrame(self.sheet.values)
        merged_cell_ranges = self.sheet.merged_cells.ranges
        if len(merged_cell_ranges) > 0:
            self.hasmerged = True
            for cell_group in merged_cell_ranges.copy():
                min_col, min_row, max_col, max_row = range_boundaries(str(cell_group))
                top_left_cell_value = self.sheet.cell(row=min_row, column=min_col).value
                self.sheet.unmerge_cells(str(cell_group))
                for row in self.sheet.iter_rows(min_col=min_col, min_row=min_row, max_col=max_col, max_row=max_row):
                    for cell in row:
                        cell.value = top_left_cell_value

        self.df = self.extract_and_clean_data()

    def extract_and_clean_data(self):
        data = []
        for row in self.sheet.iter_rows(min_row=1):
            row_data = [cell.value for cell in row]
            if any(row_data):
                data.append(row_data)

        df = pd.DataFrame(data)
        df = df.T.drop_duplicates().T
        df = df.drop_duplicates()
        return df.fillna("")

    def find_first_data_row(self):
        """
        Find the first row with data in it using simple rules
        """
        report, report_json = self.get_sheet_attributes()

        if self.model is None:
            datarow = report_json["first_not_null_row"]
            if report_json["first_float_row"] > datarow:
                datarow = report_json["first_float_row"]
            if report_json["hxl_row"] is not None:
                datarow = report_json["hxl_row"]
            if report_json["year_vals_in_rows"][datarow] > 3:
                datarow = datarow + 1
        else:
            pass # call llm agent

        return datarow, report

    def parse_excel_sheet(self):
        """
        Parses the excel sheet and returns a dataframe with the correct columns,
        collapsed merged cells, and cleaned data.
        """
        if self.debug:
            print(pd.DataFrame(self.sheet.values)[0:self.maxn])

        # Pad merged cells
        self.pad_merged_cells()

        # Identify the first row where data starts
        self.datarow, self.report = self.find_first_data_row()

        # Adjust columns if necessary for merged headers
        if self.datarow > 0 and self.hasmerged:
            self.adjust_for_merged_headers()

        # Remove duplicates and clean up the dataframe
        self.df = self.df.T.drop_duplicates().T
        self.df = self.df.fillna("")
        self.df.columns = self.df.iloc[0]
        self.df = self.df[1:]

        if self.debug:
            print(self.report)
            print(self.df.head(3))

        return self.df

    def adjust_for_merged_headers(self):
        min_row = 2
        for row in self.sheet.iter_rows(min_row=1, max_row=3):
            left_value = self.sheet.cell(row=row[0].row, column=1).value
            all_same = all(cell.value == left_value or cell.value is None for cell in row)
            if all_same:
                min_row += 1

        for row in self.sheet.iter_rows(min_row=min_row, max_row=self.datarow - 1):
            for cell in row:
                val_m1 = str(cell.offset(-1, 0).value)
                val_here = str(cell.value)
                if val_here not in val_m1:
                    cell.value = val_m1 + " - " + str(cell.value)
                else:
                    cell.value = val_m1
                cell.value = re.sub("^ +-|None", "", cell.value)
